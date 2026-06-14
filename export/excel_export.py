from typing import Optional
from pathlib import Path
from datetime import datetime
from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger
from core.settings import settings


RUSSIAN_COLUMNS: dict[str, str] = {
    "source": "Источник",
    "task_id": "ID задачи",
    "url": "Ссылка",
    "title": "Название",
    "description": "Описание",
    "budget_min": "Бюджет от",
    "budget_max": "Бюджет до",
    "currency": "Валюта",
    "posted_at": "Дата публикации",
    "proposals_count": "Отклики",
    "category_raw": "Категория (исходная)",
    "technologies": "Технологии",
    "country": "Страна",
    "client_rating": "Рейтинг заказчика",
    "payment_verified": "Оплата подтверждена",
    "normalized_category": "Категория (нормализованная)",
    "scraped_at": "Дата сбора",
}

RUSSIAN_CATEGORIES: dict[str, str] = {
    "C# / .NET": "C# / .NET",
    "Design": "Дизайн / Графика",
    "Marketing": "Маркетинг / Реклама",
    "WordPress": "WordPress",
    "Tilda": "Tilda",
    "Shopify": "Shopify",
    "Frontend": "Frontend / Вёрстка",
    "React": "React",
    "Next.js": "Next.js",
    "Backend": "Backend",
    "FastAPI": "FastAPI",
    "Django": "Django",
    "Laravel": "Laravel",
    "Fullstack": "Fullstack",
    "Telegram Bots": "Telegram-боты",
    "Discord Bots": "Discord-боты",
    "AI Chatbots": "AI-чатботы",
    "AI Agents": "AI-агенты",
    "RAG": "RAG (Retrieval Augmented Generation)",
    "OpenAI Integration": "OpenAI интеграция",
    "Claude Integration": "Claude интеграция",
    "Web Scraping": "Web Scraping / Парсинг",
    "Parsing": "Парсинг данных",
    "Automation": "Автоматизация",
    "n8n": "n8n",
    "Make": "Make (Integromat)",
    "Zapier": "Zapier",
    "Mobile Apps": "Мобильные приложения",
    "Flutter": "Flutter",
    "React Native": "React Native",
    "Android": "Android",
    "iOS": "iOS",
    "DevOps": "DevOps",
    "Docker": "Docker",
    "Kubernetes": "Kubernetes",
    "QA": "QA / Тестирование",
    "Data Analytics": "Data Analytics",
    "Power BI": "Power BI",
    "SQL": "SQL / Базы данных",
    "Machine Learning": "Machine Learning",
    "Computer Vision": "Computer Vision",
    "OTHER": "Другое (не IT)",
}

IT_CATEGORIES: set[str] = set(RUSSIAN_CATEGORIES.keys()) - {"OTHER"}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(
    start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
)


def is_it_task(task: dict) -> bool:
    cat = task.get("normalized_category", "OTHER")
    if cat in IT_CATEGORIES:
        return True
    techs = task.get("technologies")
    if techs and isinstance(techs, list) and len(techs) > 0:
        return True
    return False


def ru_category(eng: str) -> str:
    return RUSSIAN_CATEGORIES.get(eng, eng)


class ExcelExporter:
    def __init__(self) -> None:
        self.export_dir = Path(settings.export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self, tasks: list[dict], analytics: dict, filename: Optional[str] = None
    ) -> Path:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"freelance_market_analysis_{timestamp}.xlsx"

        filepath = self.export_dir / filename
        wb = Workbook()

        it_tasks = [t for t in tasks if is_it_task(t)]
        non_it_tasks = [t for t in tasks if not is_it_task(t)]
        it_count = len(it_tasks)
        non_it_count = len(non_it_tasks)

        self._write_summary(wb, it_tasks, it_count, non_it_count, analytics)
        self._write_raw_tasks(wb, it_tasks, "IT-задачи")
        self._write_top_categories(wb, analytics)
        self._write_top_technologies(wb, analytics)
        self._write_average_budget(wb, analytics)
        self._write_median_budget(wb, analytics)
        self._write_competition(wb, analytics)
        self._write_fastest_growing(wb, analytics)
        self._write_highest_paying(wb, analytics)
        self._write_examples(wb, it_tasks)

        wb.save(str(filepath))
        logger.info(
            "Excel report saved to {} ({} IT + {} non-IT tasks)",
            filepath,
            it_count,
            non_it_count,
        )
        return filepath

    @staticmethod
    def _apply_header_style(ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    @staticmethod
    def _apply_data_style(ws, start_row: int, end_row: int, col_count: int) -> None:
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if (row - start_row) % 2 == 1:
                    cell.fill = ALT_FILL

    @staticmethod
    def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
        for col_cells in ws.columns:
            column_letter = col_cells[0].column_letter
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_len = max(max_len, cell_len)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted

    @staticmethod
    def _write_styled_cell(
        ws, row: int, col: int, value, font=None, fill=None, alignment=None
    ) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        cell.border = THIN_BORDER

    def _write_summary(
        self,
        wb: Workbook,
        it_tasks: list[dict],
        it_count: int,
        non_it_count: int,
        analytics: dict,
    ) -> None:
        ws = wb.active
        ws.title = "Общая аналитика рынка"

        title_font = Font(bold=True, size=16, color="2F5496")
        subtitle_font = Font(bold=True, size=13, color="333333")
        metric_font = Font(bold=True, size=11)
        value_font = Font(size=11, color="2E7D32")
        alert_font = Font(size=11, color="C62828")
        highlight_font = Font(size=11, color="E65100")
        normal_font = Font(size=11)
        wrap_align = Alignment(wrap_text=True, vertical="top")

        row = 1
        ws.cell(row=row, column=1, value="ОБЩАЯ АНАЛИТИКА РЫНКА ФРИЛАНСА")
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        row = 2
        ws.cell(
            row=row,
            column=1,
            value=f"Дата отчёта: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        )
        ws.cell(row=row, column=1).font = Font(size=10, color="666666")

        row = 4
        ws.cell(row=row, column=1, value="СТАТИСТИКА СБОРА")
        ws.cell(row=row, column=1).font = subtitle_font

        stats = [
            ("Всего собрано задач", f"{it_count + non_it_count}", ""),
            (
                "Из них IT-задач",
                f"{it_count}",
                f"({it_count / max(it_count + non_it_count, 1) * 100:.0f}%)",
            ),
            (
                "Не IT (отфильтровано)",
                f"{non_it_count}",
                f"({non_it_count / max(it_count + non_it_count, 1) * 100:.0f}%)",
            ),
            (
                "Количество категорий",
                f"{len(analytics.get('category_counts', {}))}",
                "",
            ),
            (
                "Количество технологий",
                f"{len(analytics.get('technology_counts', {}))}",
                "",
            ),
            ("Площадки", ", ".join(analytics.get("sources", [])) or "Kwork", ""),
        ]

        for i, (label, value, extra) in enumerate(stats):
            r = row + 1 + i
            self._write_styled_cell(ws, r, 1, label, font=metric_font)
            self._write_styled_cell(
                ws,
                r,
                2,
                value,
                font=value_font
                if "IT-задач" in label or "Всего" in label
                else normal_font,
            )
            self._write_styled_cell(ws, r, 3, extra, font=normal_font)
            if i % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = SUMMARY_FILL

        row = 12
        ws.cell(row=row, column=1, value="РАСПРЕДЕЛЕНИЕ ПО КАТЕГОРИЯМ")
        ws.cell(row=row, column=1).font = subtitle_font

        cat_counts = analytics.get("category_counts", {})
        it_cat_counts = {k: v for k, v in cat_counts.items() if k in IT_CATEGORIES}
        sorted_cats = sorted(it_cat_counts.items(), key=lambda x: x[1], reverse=True)

        headers = ["Категория", "Задач", "Доля", "Статус"]
        for c, h in enumerate(headers, 1):
            self._write_styled_cell(
                ws,
                row + 1,
                c,
                h,
                font=HEADER_FONT,
                fill=HEADER_FILL,
                alignment=HEADER_ALIGNMENT,
            )

        max_count = sorted_cats[0][1] if sorted_cats else 0
        for i, (cat, count) in enumerate(sorted_cats):
            r = row + 2 + i
            share = count / max(it_count, 1) * 100
            self._write_styled_cell(ws, r, 1, ru_category(cat), font=normal_font)
            self._write_styled_cell(ws, r, 2, count, font=normal_font)
            self._write_styled_cell(ws, r, 3, f"{share:.1f}%", font=normal_font)

            if count == max_count:
                status = "Самый востребованный"
                self._write_styled_cell(
                    ws, r, 4, status, font=alert_font, fill=ALERT_FILL
                )
            elif share >= 10:
                status = "Высокий спрос"
                self._write_styled_cell(
                    ws, r, 4, status, font=highlight_font, fill=HIGHLIGHT_FILL
                )
            elif share >= 5:
                status = "Средний спрос"
                self._write_styled_cell(
                    ws, r, 4, status, font=Font(size=11, color="1565C0")
                )
            else:
                status = "Низкий спрос"
                self._write_styled_cell(
                    ws, r, 4, status, font=Font(size=11, color="9E9E9E")
                )

        row = row + 2 + len(sorted_cats) + 1
        ws.cell(row=row, column=1, value="АНАЛИЗ БЮДЖЕТОВ")
        ws.cell(row=row, column=1).font = subtitle_font

        ba = analytics.get("budget_analysis", {})
        budget_stats = [
            ("Средний бюджет", f"₽{(ba.get('average_budget_min') or 0):,.2f}"),
            ("Медианный бюджет", f"₽{(ba.get('median_budget_min') or 0):,.2f}"),
            ("Минимальный бюджет", f"₽{(ba.get('min_budget') or 0):,.2f}"),
            ("Максимальный бюджет", f"₽{(ba.get('max_budget') or 0):,.2f}"),
        ]
        for i, (label, val) in enumerate(budget_stats):
            r = row + 1 + i
            self._write_styled_cell(ws, r, 1, label, font=metric_font)
            self._write_styled_cell(ws, r, 2, val, font=value_font)

        row = row + len(budget_stats) + 2
        ws.cell(row=row, column=1, value="САМЫЕ ДОРОГИЕ КАТЕГОРИИ")
        ws.cell(row=row, column=1).font = subtitle_font

        highest_paying = analytics.get("highest_paying_categories", [])
        if highest_paying:
            for c, h in enumerate(["Категория", "Средний бюджет", "Задач"], 1):
                self._write_styled_cell(
                    ws,
                    row + 1,
                    c,
                    h,
                    font=HEADER_FONT,
                    fill=HEADER_FILL,
                    alignment=HEADER_ALIGNMENT,
                )
            for i, item in enumerate(highest_paying[:5]):
                r = row + 2 + i
                self._write_styled_cell(
                    ws, r, 1, ru_category(item["category"]), font=normal_font
                )
                self._write_styled_cell(
                    ws, r, 2, f"₽{item.get('avg_budget', 0):,.2f}", font=value_font
                )
                self._write_styled_cell(
                    ws, r, 3, item.get("task_count", 0), font=normal_font
                )

        row = row + 9
        ws.cell(row=row, column=1, value="КОНКУРЕНЦИЯ")
        ws.cell(row=row, column=1).font = subtitle_font

        ca = analytics.get("competition_analysis", {})
        comp_stats = [
            (
                "Среднее число откликов",
                f"{ca.get('overall_average_proposals') or 0:.1f}",
            ),
            ("Медиана откликов", f"{ca.get('overall_median_proposals') or 0:.1f}"),
        ]
        for i, (label, val) in enumerate(comp_stats):
            r = row + 1 + i
            self._write_styled_cell(ws, r, 1, label, font=metric_font)
            self._write_styled_cell(ws, r, 2, val, font=value_font)

        row = row + 4
        ws.cell(row=row, column=1, value="ВЫВОДЫ ПО РЫНКУ")
        ws.cell(row=row, column=1).font = subtitle_font

        conclusions = []
        if sorted_cats:
            top_cat = sorted_cats[0]
            conclusions.append(
                f"Самая востребованная категория: «{ru_category(top_cat[0])}» "
                f"({top_cat[1]} задач, {top_cat[1] / max(it_count, 1) * 100:.0f}% от всех IT-задач)."
            )

        if highest_paying:
            top_pay = highest_paying[0]
            conclusions.append(
                f"Самая высокооплачиваемая категория: «{ru_category(top_pay['category'])}» "
                f"(средний бюджет ₽{top_pay['avg_budget']:,.2f})."
            )

        if sorted_cats and len(sorted_cats) >= 2:
            least_cat = sorted_cats[-1]
            conclusions.append(
                f"Наименее востребованная категория: «{ru_category(least_cat[0])}» "
                f"({least_cat[1]} задач). Возможно, стоит присмотреться к этой нише "
                f"— низкая конкуренция."
            )

        growing = analytics.get("fastest_growing_categories", [])
        if growing:
            top_grow = growing[0]
            conclusions.append(
                f"Быстрорастущая категория: «{ru_category(top_grow['category'])}» "
                f"(рост {top_grow.get('growth_percent', 0):.1f}%)."
            )

        avg_proposals = ca.get("overall_average_proposals")
        if avg_proposals and avg_proposals > 0:
            if avg_proposals < 5:
                conclusions.append(
                    "Низкая конкуренция на рынке: в среднем "
                    f"{avg_proposals:.1f} откликов на задачу."
                )
            elif avg_proposals < 15:
                conclusions.append(
                    "Умеренная конкуренция: в среднем "
                    f"{avg_proposals:.1f} откликов на задачу."
                )
            else:
                conclusions.append(
                    "Высокая конкуренция: в среднем "
                    f"{avg_proposals:.1f} откликов на задачу. "
                    "Рекомендуется выбирать узкие ниши."
                )

        tech_counts = analytics.get("technology_counts", {})
        if tech_counts:
            top_tech = sorted(tech_counts.items(), key=lambda x: x[1], reverse=True)[0]
            conclusions.append(
                f"Самая востребованная технология: «{top_tech[0]}» "
                f"(упоминается в {top_tech[1]} задачах)."
            )

        if not conclusions:
            conclusions.append("Недостаточно данных для формирования выводов.")

        for i, conclusion in enumerate(conclusions):
            r = row + 1 + i
            self._write_styled_cell(
                ws,
                r,
                1,
                f"{i + 1}. {conclusion}",
                font=normal_font,
                alignment=wrap_align,
            )
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 25

    def _write_raw_tasks(
        self, wb: Workbook, tasks: list[dict], sheet_name: str = "IT-задачи"
    ) -> None:
        ws = wb.create_sheet(sheet_name)

        sorted_tasks = sorted(
            tasks,
            key=lambda t: (
                t.get("normalized_category", "OTHER") or "OTHER",
                t.get("title", ""),
            ),
        )

        columns = [
            "source",
            "task_id",
            "url",
            "title",
            "description",
            "budget_min",
            "budget_max",
            "currency",
            "posted_at",
            "proposals_count",
            "category_raw",
            "technologies",
            "country",
            "client_rating",
            "payment_verified",
            "normalized_category",
            "scraped_at",
        ]

        for col_idx, col_name in enumerate(columns, 1):
            ru_name = RUSSIAN_COLUMNS.get(col_name, col_name)
            ws.cell(row=1, column=col_idx, value=ru_name)

        self._apply_header_style(ws, 1, len(columns))

        for row_idx, task in enumerate(sorted_tasks, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = task.get(col_name)
                if col_name == "normalized_category":
                    value = ru_category(str(value)) if value else "Другое"
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                elif isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, bool):
                    value = "Да" if value else "Нет"
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._apply_data_style(ws, 2, len(tasks) + 1, len(columns))
        self._auto_width(ws)

        last_col = ws.cell(row=1, column=len(columns)).column_letter
        last_row = len(tasks) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"

    def _write_top_categories(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Топ категорий")
        category_counts = analytics.get("category_counts", {})
        it_cat_counts = {k: v for k, v in category_counts.items() if k in IT_CATEGORIES}

        headers = ["Категория", "Количество задач", "Доля (%)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        total = sum(it_cat_counts.values()) or 1
        sorted_cats = sorted(it_cat_counts.items(), key=lambda x: x[1], reverse=True)

        for row_idx, (cat, count) in enumerate(sorted_cats, 2):
            ws.cell(row=row_idx, column=1, value=ru_category(cat))
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=round(count / total * 100, 2))

        self._apply_data_style(ws, 2, len(sorted_cats) + 1, len(headers))
        self._auto_width(ws)

    def _write_top_technologies(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Топ технологий")
        tech_counts = analytics.get("technology_counts", {})

        headers = ["Технология", "Упоминания", "Доля (%)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        total = sum(tech_counts.values()) or 1
        sorted_techs = sorted(tech_counts.items(), key=lambda x: x[1], reverse=True)

        for row_idx, (tech, count) in enumerate(sorted_techs, 2):
            ws.cell(row=row_idx, column=1, value=tech)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=round(count / total * 100, 2))

        self._apply_data_style(ws, 2, len(sorted_techs) + 1, len(headers))
        self._auto_width(ws)

    def _write_average_budget(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Средний бюджет")
        budget_analysis = analytics.get("budget_analysis", {})

        ws.append(["Показатель", "Значение"])
        self._apply_header_style(ws, 1, 2)

        metrics = [
            ("Средний бюджет (мин)", budget_analysis.get("average_budget_min")),
            ("Средний бюджет (макс)", budget_analysis.get("average_budget_max")),
            ("Медианный бюджет (мин)", budget_analysis.get("median_budget_min")),
            ("Медианный бюджет (макс)", budget_analysis.get("median_budget_max")),
            ("Минимальный бюджет", budget_analysis.get("min_budget")),
            ("Максимальный бюджет", budget_analysis.get("max_budget")),
            ("Стд. отклонение (мин)", budget_analysis.get("std_dev_budget_min")),
            ("Стд. отклонение (макс)", budget_analysis.get("std_dev_budget_max")),
            ("Задач с бюджетом", budget_analysis.get("total_tasks_with_budget")),
            ("Задач без бюджета", budget_analysis.get("total_tasks_without_budget")),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 2):
            ws.cell(row=row_idx, column=1, value=metric)
            if value is not None:
                ws.cell(row=row_idx, column=2, value=round(value, 2))
            else:
                ws.cell(row=row_idx, column=2, value="Н/Д")

        self._apply_data_style(ws, 2, len(metrics) + 1, 2)

        cat_avg = budget_analysis.get("category_average_budgets", {})
        if cat_avg:
            start_row = len(metrics) + 3
            ws.cell(row=start_row, column=1, value="По категориям")
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
            start_row += 1
            ws.cell(row=start_row, column=1, value="Категория")
            ws.cell(row=start_row, column=2, value="Средний бюджет")
            self._apply_header_style(ws, start_row, 2)

            sorted_cats = sorted(cat_avg.items(), key=lambda x: x[1], reverse=True)
            for i, (cat, avg) in enumerate(sorted_cats, start_row + 1):
                ws.cell(row=i, column=1, value=ru_category(cat))
                ws.cell(row=i, column=2, value=round(avg, 2))

        self._auto_width(ws)

    def _write_median_budget(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Медианный бюджет")
        budget_analysis = analytics.get("budget_analysis", {})
        cat_median = budget_analysis.get("category_median_budgets", {})

        headers = ["Категория", "Медианный бюджет"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        sorted_cats = sorted(cat_median.items(), key=lambda x: x[1], reverse=True)
        for row_idx, (cat, med) in enumerate(sorted_cats, 2):
            ws.cell(row=row_idx, column=1, value=ru_category(cat))
            ws.cell(row=row_idx, column=2, value=round(med, 2))

        self._apply_data_style(ws, 2, len(sorted_cats) + 1, len(headers))
        self._auto_width(ws)

    def _write_competition(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Конкуренция")
        competition = analytics.get("competition_analysis", {})

        headers = ["Категория", "Среднее число откликов"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        avg_proposals = competition.get("average_proposals_per_category", {})
        sorted_cats = sorted(avg_proposals.items(), key=lambda x: x[1], reverse=True)

        for row_idx, (cat, avg) in enumerate(sorted_cats, 2):
            ws.cell(row=row_idx, column=1, value=ru_category(cat))
            ws.cell(row=row_idx, column=2, value=round(avg, 2))

        self._apply_data_style(ws, 2, len(sorted_cats) + 1, len(headers))
        self._auto_width(ws)

    def _write_fastest_growing(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Быстрорастущие категории")
        fastest_growing = analytics.get("fastest_growing_categories", [])

        headers = ["Категория", "Всего задач", "Темп роста", "Рост (%)"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        for row_idx, item in enumerate(fastest_growing, 2):
            ws.cell(row=row_idx, column=1, value=ru_category(item.get("category", "")))
            ws.cell(row=row_idx, column=2, value=item.get("total_count"))
            ws.cell(row=row_idx, column=3, value=round(item.get("growth_rate", 0), 4))
            ws.cell(
                row=row_idx, column=4, value=round(item.get("growth_percent", 0), 2)
            )

        self._apply_data_style(ws, 2, len(fastest_growing) + 1, len(headers))
        self._auto_width(ws)

    def _write_highest_paying(self, wb: Workbook, analytics: dict) -> None:
        ws = wb.create_sheet("Самые дорогие категории")
        highest_paying = analytics.get("highest_paying_categories", [])

        headers = ["Категория", "Средний бюджет", "Медиана", "Макс.", "Мин.", "Задач"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        for row_idx, item in enumerate(highest_paying, 2):
            ws.cell(row=row_idx, column=1, value=ru_category(item.get("category", "")))
            ws.cell(row=row_idx, column=2, value=round(item.get("avg_budget", 0), 2))
            ws.cell(row=row_idx, column=3, value=round(item.get("median_budget", 0), 2))
            ws.cell(row=row_idx, column=4, value=round(item.get("max_budget", 0), 2))
            ws.cell(row=row_idx, column=5, value=round(item.get("min_budget", 0), 2))
            ws.cell(row=row_idx, column=6, value=item.get("task_count"))

        self._apply_data_style(ws, 2, len(highest_paying) + 1, len(headers))
        self._auto_width(ws)

    def _write_examples(self, wb: Workbook, tasks: list[dict]) -> None:
        ws = wb.create_sheet("Примеры задач")
        headers = ["Категория", "Название", "Бюджет", "Источник", "Ссылка"]
        ws.append(headers)
        self._apply_header_style(ws, 1, len(headers))

        by_category: dict[str, list] = {}
        for task in tasks:
            cat = task.get("normalized_category", "OTHER")
            if cat in IT_CATEGORIES:
                if cat not in by_category:
                    by_category[cat] = []
                if len(by_category[cat]) < 5:
                    by_category[cat].append(task)

        row_idx = 2
        for category in sorted(by_category.keys()):
            for task in by_category[category]:
                budget = task.get("budget_min") or task.get("budget_max")
                budget_str = f"₽{budget:,}" if budget else "Н/Д"

                ws.cell(row=row_idx, column=1, value=ru_category(category))
                ws.cell(row=row_idx, column=2, value=task.get("title"))
                ws.cell(row=row_idx, column=3, value=budget_str)
                ws.cell(row=row_idx, column=4, value=task.get("source"))
                ws.cell(row=row_idx, column=5, value=task.get("url"))
                row_idx += 1

        self._apply_data_style(ws, 2, row_idx - 1, len(headers))
        self._auto_width(ws)
